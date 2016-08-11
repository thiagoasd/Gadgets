import re
import sys
from os import listdir;
from os.path import isdir, join;
from openpyxl import Workbook;
from openpyxl.styles import PatternFill, Alignment, Font


def header(sheet):
    font = Font(bold=True)
    align = Alignment(horizontal="center")
    sheet["A1"] = "CODE"
    sheet["B1"] = "ALBUM NAME"
    sheet["C1"] = "CONTENT TYPE"
    sheet["D1"] = "LOSSLESS"
    sheet["E1"] = "WAV"
    sheet["F1"] = "CUE"

    row = sheet["A1:F1"]
    for tup in row:
        for cell in tup:
            cell.font = font
            cell.alignment = align
    return;


def filler(sheet, row, code, album, content, lossless, wav, cue):
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')

    greenFill = PatternFill(start_color='08FC00',
                            end_color='08FC00',
                            fill_type='solid')

    cell = sheet.cell(row=row, column=1)
    cell.alignment = Alignment(horizontal="center")
    cell.value = code;

    cell = sheet.cell(row=row, column=2)
    cell.alignment = Alignment(horizontal="center")
    cell.value = album;

    cell = sheet.cell(row=row, column=3)
    cell.alignment = Alignment(horizontal="center")
    cell.value = content;

    cell = sheet.cell(row=row, column=4)
    cell.alignment = Alignment(horizontal="center")
    cell.value = lossless;
    if lossless:
        cell.fill = greenFill
    else:
        cell.fill = redFill

    cell = sheet.cell(row=row, column=5)
    cell.alignment = Alignment(horizontal="center")
    cell.value = wav;
    if wav:
        cell.fill = greenFill
    else:
        cell.fill = redFill

    cell = sheet.cell(row=row, column=6)
    cell.alignment = Alignment(horizontal="center")
    cell.value = cue;
    if cue:
        cell.fill = greenFill
    else:
        cell.fill = redFill


    return;


path = str(sys.argv[1]);
book = Workbook();
sheet = book.active;
header(sheet);

row_number = 2;
for f in listdir(path):
    column_number = 1;
    cpath = join(path, f);
    if isdir(cpath):
        m = re.search('\[(.*)\](.*)\[(.*)\]', cpath);
        if m is None or len(m.groups()) < 3:
            print("Error " + cpath);
            continue;
        # CODE
        code = m.group(1);
        # ALBUM NAME
        album = m.group(2);
        # CONTENT TYPE
        content = m.group(3);

        lossless = "flac" in content or "wav" in content;
        wav = "wav" in content;
        cue = "cue" in content;

        filler(sheet, row_number, code, album, content, lossless, wav, cue)
        row_number += 1;

book.save(path + "\\test.xlsx");

# TODO GET APE, TRANSCODE FILES
